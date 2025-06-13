import openpyxl
import requests
import time

# API endpoint
API_URL = "Add the API Endpoint here."

# Translation function using confirmed working format
def translate_text(text, source_lang="en", target_lang="ne"):
    if not text.strip():
        return text

    payload = {
        "input": [{"source": text}],
        "config": {
            "language": {
                "sourceLanguage": source_lang,
                "targetLanguage": target_lang
            },
            "model": " "
        }
    }

    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json"
    }

    try:
        response = requests.post(API_URL, headers=headers, json=payload)
        if response.status_code in [200, 201]:
            result = response.json()
            return result['output'][0]['target']
        else:
            print(f"Error {response.status_code} for '{text}': {response.text}")
            return text
    except Exception as e:
        print(f"Exception while translating '{text}': {e}")
        return text

# Excel processor
def translate_excel_file(file_path, source_lang="en", target_lang="ne"):
    wb = openpyxl.load_workbook(file_path)

    for sheet_name in wb.sheetnames:
        original_sheet = wb[sheet_name]
        translated_sheet = wb.copy_worksheet(original_sheet)
        translated_sheet.title = f"{sheet_name}_translated"

        print(f"\nTranslating sheet: {sheet_name} → {translated_sheet.title}")

        for row in translated_sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip():
                    original = cell.value
                    translated = translate_text(str(original), source_lang, target_lang)
                    cell.value = translated
                    # time.sleep(0.3)  

    # Save output
    output_file = "translated_" + file_path
    wb.save(output_file)
    print(f"\nSaved translated workbook to: {output_file}")

# Run script
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Translate text in Excel sheets using CFILT API.")
    parser.add_argument("--filepath", help="Path to the Excel file to translate (e.g., input.xlsx)")
    parser.add_argument("--source", default="en", help="Source language code (default: en)")
    parser.add_argument("--target", default="ne", help="Target language code (default: ne)")

    args = parser.parse_args()
    translate_excel_file(args.filename, source_lang=args.source, target_lang=args.target)
